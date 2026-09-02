from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import AreaChart, BarChart, LineChart, PieChart, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from config import load_config


def export_heritage_to_excel(data: dict[str, Any], currency: str) -> None:
    """Génère un classeur Excel structuré avec les 4 graphiques disposés en grille 2x2."""
    wb = Workbook()

    # Palette de styles et couleurs
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    title_font = Font(name="Calibri", size=16, bold=True, color="004D40")
    section_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    bold_font = Font(name="Calibri", size=11, bold=True)

    # Remplissages
    title_fill = PatternFill(start_color="E0F2F1", end_color="E0F2F1", fill_type="solid")
    header_fill = PatternFill(start_color="00796B", end_color="00796B", fill_type="solid")
    kpi_fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")

    # Bordures
    border_thin = Side(border_style="thin", color="B0BEC5")
    cell_border = Border(left=border_thin, right=border_thin, top=border_thin, bottom=border_thin)

    # Onglet 2 : Historique des données brutes
    ws_history = wb.active
    ws_history.title = "Historique"
    ws_history.views.sheetView[0].showGridLines = True

    history_df = data["all_accounts_df"].copy()
    history_df["Patrimoine Total"] = data["heritage_series"]
    history_df.index = history_df.index.strftime("%Y-%m-%d")

    ws_history.cell(row=1, column=1, value="Date").font = header_font
    ws_history.cell(row=1, column=1).fill = header_fill

    for col_idx, col_name in enumerate(history_df.columns, start=2):
        cell = ws_history.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for r_idx, (date_str, row) in enumerate(history_df.iterrows(), start=2):
        ws_history.cell(row=r_idx, column=1, value=date_str)
        for c_idx, val in enumerate(row, start=2):
            cell = ws_history.cell(row=r_idx, column=c_idx, value=None if pd.isna(val) else val)
            cell.number_format = f"#,##0.00 {currency}"

    max_row_hist = len(history_df) + 1
    max_col_hist = len(history_df.columns) + 1
    total_col_idx = max_col_hist

    # Onglet 1 : Tableau de Bord
    ws_dash = wb.create_sheet(title="Tableau de bord", index=0)
    ws_dash.views.sheetView[0].showGridLines = True

    # Banner Titre Principal
    ws_dash.merge_cells("A1:E1")
    title_cell = ws_dash["A1"]
    title_cell.value = "TABLEAU DE BORD PATRIMONIAL"
    title_cell.font = title_font
    title_cell.fill = title_fill
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws_dash.row_dimensions[1].height = 40

    # Section 1 : Indicateur Clé (KPI Total Patrimoine)
    ws_dash.merge_cells("A3:E3")
    kpi_header = ws_dash["A3"]
    kpi_header.value = "INDICATEUR CLÉ"
    kpi_header.font = section_font
    kpi_header.fill = header_fill
    kpi_header.alignment = Alignment(horizontal="center")

    ws_dash.merge_cells("A4:B4")
    ws_dash["A4"] = "Patrimoine Total Actuel"
    ws_dash["A4"].font = bold_font
    ws_dash["A4"].alignment = Alignment(horizontal="center")

    total_amount = data["heritage_series"].iloc[-1] if not data["heritage_series"].empty else 0.0
    ws_dash.merge_cells("C4:E4")
    total_cell = ws_dash["C4"]
    total_cell.value = total_amount
    total_cell.font = Font(name="Calibri", size=14, bold=True, color="004D40")
    total_cell.number_format = f"#,##0.00 {currency}"
    total_cell.alignment = Alignment(horizontal="center")

    for r in range(3, 5):
        for c in range(1, 6):
            cell = ws_dash.cell(row=r, column=c)
            cell.border = cell_border
            if r == 4 and c < 3:
                cell.fill = kpi_fill

    # Section 2 : Tables (Répartition à gauche / Évolution Annuelle à droite)
    ws_dash.merge_cells("A6:C6")
    ws_dash["A6"] = "RÉPARTITION PAR COMPTE"
    ws_dash["A6"].font = section_font
    ws_dash["A6"].fill = header_fill
    ws_dash["A6"].alignment = Alignment(horizontal="center")

    ws_dash.merge_cells("D6:E6")
    ws_dash["D6"] = "ÉVOLUTION ANNUELLE"
    ws_dash["D6"].font = section_font
    ws_dash["D6"].fill = header_fill
    ws_dash["D6"].alignment = Alignment(horizontal="center")

    # En-têtes Répartition
    headers_rep = ["Compte", "Solde", "Part"]
    for i, h in enumerate(headers_rep, start=1):
        c = ws_dash.cell(row=7, column=i, value=h)
        c.font = bold_font
        c.fill = kpi_fill
        c.alignment = Alignment(horizontal="center")
        c.border = cell_border

    # En-têtes Performance Annuelle
    headers_perf = ["Année", "Variation"]
    for i, h in enumerate(headers_perf, start=4):
        c = ws_dash.cell(row=7, column=i, value=h)
        c.font = bold_font
        c.fill = kpi_fill
        c.alignment = Alignment(horizontal="center")
        c.border = cell_border

    # Remplissage Table Répartition
    row_rep = 8
    for account_name, details in data["account_distribution"].items():
        c1 = ws_dash.cell(row=row_rep, column=1, value=account_name)
        c2 = ws_dash.cell(row=row_rep, column=2, value=details["amount"])
        c3 = ws_dash.cell(row=row_rep, column=3, value=details["percentage"] / 100)

        c2.number_format = f"#,##0.00 {currency}"
        c3.number_format = "0.00%"

        for cell in (c1, c2, c3):
            cell.border = cell_border

        row_rep += 1

    # Remplissage Table Performance Annuelle
    row_perf = 8
    yearly_df = data.get("yearly_growth")
    if yearly_df is not None and not yearly_df.empty:
        for year, row in yearly_df.iterrows():
            pct_val = row["percentage_change"] / 100
            c1 = ws_dash.cell(row=row_perf, column=4, value=int(year))
            c2 = ws_dash.cell(row=row_perf, column=5, value=pct_val)

            c1.alignment = Alignment(horizontal="center")
            c2.number_format = "+0.00%;-0.00%;0.00%"

            text_color = "008000" if pct_val >= 0 else "FF0000"
            c2.font = Font(name="Calibri", size=11, bold=True, color=text_color)

            for cell in (c1, c2):
                cell.border = cell_border
            row_perf += 1

    # Disposition des graphiques en carrés
    dates_ref = Reference(ws_history, min_col=1, min_row=2, max_row=max_row_hist)

    # Haut-Gauche : Évolution du Patrimoine Global
    area = AreaChart()
    area.title = "Évolution du Patrimoine Global"
    area.style = 13
    area.y_axis.title = f"Montant ({currency})"
    area.x_axis.title = "Date"
    data_total = Reference(ws_history, min_col=total_col_idx, min_row=1, max_row=max_row_hist)
    area.add_data(data_total, titles_from_data=True)
    area.set_categories(dates_ref)
    area.width = 15
    area.height = 9.5
    ws_dash.add_chart(area, "G1")

    # Haut-Droite : Évolution Détaillée par Compte
    line = LineChart()
    line.title = "Évolution détaillée par compte"
    line.y_axis.title = f"Montant ({currency})"
    line.x_axis.title = "Date"

    # Masquer la légende s'il n'y a qu'un seul compte
    num_accounts = len(data["all_accounts_df"].columns)
    if num_accounts <= 1:
        line.legend = None

    data_accounts = Reference(ws_history, min_col=2, min_row=1, max_col=total_col_idx - 1, max_row=max_row_hist)
    line.add_data(data_accounts, titles_from_data=True)
    line.set_categories(dates_ref)
    line.width = 15
    line.height = 9.5
    ws_dash.add_chart(line, "P1")

    # Bas-Gauche : Répartition par compte
    pie = PieChart()
    pie.title = "Répartition du patrimoine par compte"
    labels_pie = Reference(ws_dash, min_col=1, min_row=8, max_row=row_rep - 1)
    data_pie = Reference(ws_dash, min_col=2, min_row=7, max_row=row_rep - 1)
    pie.add_data(data_pie, titles_from_data=True)
    pie.set_categories(labels_pie)
    pie.width = 15
    pie.height = 9.5
    ws_dash.add_chart(pie, "G17")

    # Bas-Droite : Performance Annuelle
    if yearly_df is not None and not yearly_df.empty:
        bar = BarChart()
        bar.type = "col"
        bar.title = "Performance annuelle du patrimoine (%)"
        bar.y_axis.title = "Variation (%)"
        bar.x_axis.title = "Année"
        bar.legend = None

        data_bar = Reference(ws_dash, min_col=5, min_row=7, max_row=row_perf - 1)
        labels_bar = Reference(ws_dash, min_col=4, min_row=8, max_row=row_perf - 1)
        bar.add_data(data_bar, titles_from_data=True)
        bar.set_categories(labels_bar)
        bar.width = 15
        bar.height = 9.5

        # Attribution des couleurs (Vert / Rouge) par barre
        series = bar.series[0]
        for idx, (_, row) in enumerate(yearly_df.iterrows()):
            val = row["percentage_change"]
            color_hex = "008000" if val >= 0 else "FF0000"

            dp = DataPoint(idx=idx)
            dp.graphicalProperties.solidFill = color_hex
            series.dPt.append(dp)

        ws_dash.add_chart(bar, "P17")

    # Auto-fit de la largeur des colonnes
    for ws in [ws_dash, ws_history]:
        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 14)

    output_path = Path(load_config()["destination_path"]) / "heritage" / "heritage_global.xlsx"
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
